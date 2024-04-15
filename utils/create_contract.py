import random
import requests
from loguru import logger
import string
from web3 import Web3
from utils.wallet import Wallet
from utils.retry import exception_handler
import json as js
import openpyxl
import os


class CreateContract(Wallet):

    def __init__(self, private_key, chain, number, proxy):
        super().__init__(private_key, chain, number, proxy)
        self.address = Web3.to_checksum_address('0x777777C338d93e2C7adf08D102d45CA7CC4Ed021')
        self.abi = js.load(open('./abi/create_contract.txt'))
        self.contract = self.web3.eth.contract(address=self.address, abi=self.abi)

    @staticmethod
    def upload_image():
        picture_list = os.listdir('picture')
        file_list = []
        for picture in picture_list:
            with open(f'picture/{picture}', "rb") as f:
                file = f.read()
                file_list.append(file)
        pic = random.choice(file_list)

        files = {"file": pic}
        url = 'https://ipfs-uploader.zora.co/api/v0/add?stream-channels=true&cid-version=1&progress=false'
        res = requests.post(url=url, files=files, timeout=60)
        json_data = res.json()
        return 'ipfs://' + json_data['Hash']

    @staticmethod
    def get_name():
        characters = string.ascii_letters + string.digits  # буквы и цифры
        random_string = ''.join(random.choice(characters) for _ in range(7))
        return random_string

    @exception_handler('Create NFT Zora.co')
    def create_contarct(self):
        logger.info('Create NFT Zora.co')

        dick = {
            'from': self.address_wallet,
            'nonce': self.web3.eth.get_transaction_count(self.address_wallet),
            **self.get_gas_price()
        }

        name = 'Stalker' + str(random.randint(1, 1000))

        txn = self.contract.functions.createContract(
            self.upload_image(),
            name,
            [0, 500, self.address_wallet],
            self.address_wallet,

            [
                "0xe72878b40000000000000000000000000000000000000000000000000000000000000000",
                "0x674cbae60000000000000000000000000000000000000000000000000000000000000060000000000000000000000000000000000000000000000000ffffffffffffffff00000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000042697066733a2f2f6261666b72656966716877766a7535626d77683562336871717a6c3636336473707a32786b7568646965723462343737323533323676786a657734000000000000000000000000000000000000000000000000000000000000",
                "0xafed7e9e0000000000000000000000000000000000000000000000000000000000000001000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000001f4000000000000000000000000" + self.address_wallet[2:],
                "0x8ec998a00000000000000000000000000000000000000000000000000000000000000001000000000000000000000000777777e8850d8d6d98de2b5f64fae401f96eff310000000000000000000000000000000000000000000000000000000000000004",
                "0xd904b94a0000000000000000000000000000000000000000000000000000000000000001000000000000000000000000777777e8850d8d6d98de2b5f64fae401f96eff31000000000000000000000000000000000000000000000000000000000000006000000000000000000000000000000000000000000000000000000000000000e4ce35b8af000000000000000000000000000000000000000000000000000000000000000100000000000000000000000000000000000000000000000000000000661cdf150000000000000000000000000000000000000000000000000000000066446c15000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000000038d7ea4c68000000000000000000000000000" + self.address_wallet[2:] + "000000000000000000000000078540eecc8b6d89949c9c7d5e8e91eab64f669600000000000000000000000000000000000000000000000000000000",
                "0xc238d1ee000000000000000000000000" + self.address_wallet[2:] + "00000000000000000000000000000000000000000000000000000000000000010000000000000000000000000000000000000000000000000000000000000001000000000000000000000000000000000000000000000000000000000000008000000000000000000000000000000000000000000000000000000000000000140000000000000000000000000000000000000000000000000000000000000000"
            ]
        ).build_transaction(dick)

        tx_hash = self.send_transaction_and_wait(txn, f'Create contract NFT')
        addr = self.save_address(tx_hash)
        self.write_address(addr)

    @exception_handler()
    def save_address(self, tx_hash):

        txn_receipt = self.web3.eth.get_transaction_receipt(tx_hash)
        addres_nft = txn_receipt.logs[0].address

        with open('./data/address_nft.txt', 'a', encoding='utf-8') as f:
            f.write(f'{addres_nft}\n')

        return addres_nft

    @exception_handler()
    def write_address(self, address):
        workbook = openpyxl.load_workbook('./data/accounts_data.xlsx')

        sheet = workbook.active

        for row_number, row in enumerate(sheet.iter_rows(min_row=1, max_col=1, max_row=sheet.max_row, values_only=True), start=1):
            if row[0] == self.private_key:

                sheet.cell(row=row_number, column=3, value=address)

                workbook.save('./data/accounts_data.xlsx')

                logger.success(f"Data successfully written to address {address}\n")
                return
